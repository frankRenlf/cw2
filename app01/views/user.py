from django.shortcuts import render, redirect
from app01 import models
from app01.utils.Pagination import Pagination
from app01.modelForms.UserModelForm import UserModelForm
from datetime import datetime

""" create userinfo operations """


def user_list(request):
    user_union = models.UserInfo.objects.all()
    pagination = Pagination(request, user_union)
    return render(request, 'user_list.html', {'user_union': pagination.number_list, "page_list": pagination.page_list})


def user_add(request):
    if request.method == "GET":
        departs = models.Department.objects.all()
        gender_choices = models.UserInfo.gender_choices
        return render(request, 'user_add.html', {"departs": departs, "gender_choices": gender_choices})
    user = models.UserInfo
    user.name = request.POST.get("name")
    user.password = request.POST.get("password")
    user.age = request.POST.get("age")
    user.gender = request.POST.get("gender")
    user.salary = request.POST.get("salary")
    user.create_time = request.POST.get("create_time")
    user.depart.id = request.POST.get("depart_id")
    models.UserInfo.objects.create(name=user.name, password=user.password, age=user.age, gender=user.gender,
                                   salary=user.salary,
                                   create_time=user.create_time,
                                   depart_id=user.depart.id)
    return redirect('/user/list')


def user_modelform_add(request):
    if request.method == "GET":
        user_null = UserModelForm()
        return render(request, 'userForm_add.html', {"userForm": user_null})
    user_form = UserModelForm(data=request.POST)
    if user_form.is_valid():
        user_form.save()
        return redirect('/user/list')
    return render(request, 'userForm_add.html', {"userForm": user_form})


def user_modelform_edit(request, uid):
    user = models.UserInfo.objects.filter(id=uid).first()
    if request.method == "GET":
        user_null = UserModelForm(instance=user)
        return render(request, 'userForm_edit.html', {"userForm": user_null})
    user_form = UserModelForm(data=request.POST, instance=user)
    if user_form.is_valid():
        user_form.save()
        return redirect('/user/list')
    return render(request, 'userForm_edit.html', {"userForm": user_form})
    # return HttpResponse(uid)


def user_delete(request):
    uid = request.GET.get("id")
    models.UserInfo.objects.filter(id=uid).delete()
    return redirect('/user/list')


def user_upload(request):
    # acquire file
    file_obj = request.FILES.get("file")
    if file_obj is None:
        return redirect('/user/list')

    # convey to openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        models.UserInfo.objects.create(name=row[0].value, password=row[1].value, age=row[2].value,
                                       gender=row[3].value, salary=row[4].value,
                                       create_time=datetime.now(), depart_id=row[6].value)
    return redirect('/user/list')
